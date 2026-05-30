"""Workbook export — v2's port of v1's ``ly_tables.tbl_workbook`` / ``tbl_sheet`` (Phase 9).

Build .xlsx workbooks from a screen's :class:`WorkbookExport` config. The endpoint splits
the screen's read query into N groups by the configured ``split_by`` column (one xlsx per
group), runs each sheet's query with the group key bound, writes the rows into an openpyxl
workbook, and streams the result. Single group ⇒ a single .xlsx; several groups ⇒ a .zip
of files.

Why server-side rather than browser-side: large datasets (a per-department LDAP report can
easily produce 100k+ rows across all sheets) can't reasonably round-trip through the
browser's memory; the SQL also needs to run on the screen's pool, so the relevant
``writable`` / row-cap / dictionary-rule machinery already lives here. The frontend just
fires the endpoint, sees the ``Content-Disposition`` filename, and triggers a download.
"""

from __future__ import annotations

import io
import re
import zipfile
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from liberty.auth.dependencies import CurrentPrincipal
from liberty.auth.principal import Principal
from liberty.connectors import ConnectorRegistry
from liberty.connectors.base import ConnectorError
from liberty.screens import Screen, ScreensFile
from liberty.screens.config import SheetSpec, WorkbookExport
from liberty.web.deps import get_connectors, get_screens, request_language
from liberty.web.errors import http_for_connector_error

router = APIRouter(prefix="/api", tags=["export"])

Screens = Annotated[ScreensFile, Depends(get_screens)]
Connectors = Annotated[ConnectorRegistry, Depends(get_connectors)]


# Excel sheet-name rules: max 31 chars; ``: \ / ? * [ ]`` forbidden + no surrounding apostrophes.
_FORBIDDEN_SHEET_CHARS = re.compile(r"[:\\/\?\*\[\]]")


def _safe_sheet_name(name: str, taken: set[str]) -> str:
    """Excel's sheet-name constraints are tighter than what a SheetSpec.name can carry —
    sanitise + dedup. Trims to 31 chars, replaces illegal chars with ``_``, suffixes
    ``_2``/``_3``/… on collision so each tab is uniquely addressable."""
    base = _FORBIDDEN_SHEET_CHARS.sub("_", (name or "Sheet").strip()) or "Sheet"
    base = base[:31]
    candidate = base
    n = 1
    while candidate in taken:
        n += 1
        suffix = f"_{n}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
    taken.add(candidate)
    return candidate


def _substitute(template: str | None, ctx: dict[str, Any]) -> str:
    """Replace ``{{key}}`` occurrences with ``ctx[key]`` (string-cast). Same shape the API
    connector + action runner use for placeholders. Used here for sheet names + file names
    so the operator's ``"{{split_value}} apps"`` becomes a real per-group sheet title."""
    out = template or ""
    for k, v in ctx.items():
        token = "{{" + k + "}}"
        if token in out:
            out = out.replace(token, "" if v is None else str(v))
    return out


def _bind_params(sheet: SheetSpec, ctx: dict[str, Any]) -> dict[str, Any]:
    """Resolve a sheet's ParamBinds against the running export context (today: just
    ``split_value`` — the parent row's group key). Same ParamBind shape used everywhere; the
    runtime here is simple enough to inline rather than reach for the action-runner's
    ``resolveBinds``.

    Result is a plain ``dict[str, str | None]`` ready to hand to ``SQLConnector.execute``.
    Empty / missing source resolves through the bind's ``default``, then drops if still empty
    (matches the action runner's convention: an unbound :param stays SQL NULL).
    """
    out: dict[str, Any] = {}
    for b in sheet.param_binds:
        # Mode A — literal ``value``. Wins outright (matches actionRunner's same-mode rule).
        if b.value is not None and b.value != "":
            out[b.param] = b.value
            continue
        # Mode B — read from the export context. The only key the export layer exposes today
        # is ``split_value``; future expansion (parent-row columns, built-ins) goes here.
        if b.source:
            v = ctx.get(b.source)
            if v is not None and str(v) != "":
                out[b.param] = v
                continue
        # Neither resolved → fall back to the bind's ``default``. Same semantics the action
        # runner uses (default is bound only when source resolves to NULL / empty).
        if b.default not in (None, ""):
            out[b.param] = b.default
    return out


def _cell_value(v: Any) -> Any:
    """Coerce a SQL cell into something openpyxl will write cleanly. Strings get the
    ``ILLEGAL_CHARACTERS_RE`` stripping openpyxl would otherwise raise on; everything else
    passes through. ``None`` becomes empty (openpyxl writes blank cells)."""
    if v is None:
        return None
    if isinstance(v, str):
        return ILLEGAL_CHARACTERS_RE.sub("", v)
    return v


async def _build_one_workbook(
    *,
    spec: WorkbookExport,
    split_value: Any,
    screen: Screen,
    app: str,
    connectors: ConnectorRegistry,
    language: str | None,
    user: str | None,
) -> bytes:
    """Build one xlsx (as ``bytes``) — one sheet per :class:`SheetSpec`. Each sheet's query
    runs with ``split_value`` bound through the sheet's ``param_binds``. Sheets are written
    in declaration order; the openpyxl ``Workbook`` is materialised in-memory and the bytes
    returned for the caller to either send raw or stuff into a zip."""
    wb = Workbook()
    # ``Workbook()`` ships with a default "Sheet" — remove it so we don't carry an empty
    # tab into the output.
    if wb.active is not None:
        wb.remove(wb.active)

    ctx: dict[str, Any] = {"split_value": split_value, "screen": screen.id, "app": app}
    taken_sheet_names: set[str] = set()
    for sheet in spec.sheets:
        conn_name = sheet.connector or screen.connector or app
        try:
            conn = connectors.sql(conn_name)
        except ConnectorError as exc:
            raise http_for_connector_error(exc) from exc
        try:
            result = await conn.execute(
                sheet.query, _bind_params(sheet, ctx),
                language=language, user=user,
            )
        except ConnectorError as exc:
            raise http_for_connector_error(exc) from exc

        # Header row + body shared by every worksheet emitted from this query. ``result.rows``
        # is keyed by the discovered (cursor.description) name; pull values by that order so
        # the header line and data lines stay column-aligned. Labels use the dictionary's
        # resolved display titles when present, same convention as TableView's CSV/XLSX export.
        col_names = [c.name for c in result.columns]
        col_labels = [c.label or c.name for c in result.columns]

        if sheet.split_by:
            # **Fan-out**: partition this query's rows by ``split_by`` into N worksheets.
            # Case-insensitive column match (Postgres folds unquoted identifiers; the operator
            # may type the column in any case). One DB roundtrip per ``SheetSpec`` either way
            # — the partition is purely in-memory.
            target = sheet.split_by.upper()
            name_map = {n.upper(): n for n in col_names}
            real_name = name_map.get(target)
            if real_name is None:
                raise HTTPException(
                    status.HTTP_422_UNPROCESSABLE_CONTENT,
                    detail=(
                        f"sheet split_by column {sheet.split_by!r} not in query "
                        f"{sheet.query!r}'s result columns"
                    ),
                )
            # Preserve first-seen order so the sheet sequence is deterministic + matches what
            # the operator sees in the TableView. Empty / None values share a single bucket.
            buckets: dict[Any, list[dict[str, Any]]] = {}
            for row in result.rows:
                rv = row.get(real_name)
                key = rv if (rv is not None and rv != "") else None
                buckets.setdefault(key, []).append(row)
            if not buckets:
                # No rows at all — emit a single empty worksheet so the operator can see the
                # sheet exists (no rows but headers). Use ``{{sheet_value}}`` → empty string.
                buckets = {None: []}
            for sheet_value, rows in buckets.items():
                sub_ctx = {**ctx, "sheet_value": sheet_value}
                sheet_name = _safe_sheet_name(
                    _substitute(sheet.name, sub_ctx) or sheet.query, taken_sheet_names,
                )
                ws = wb.create_sheet(title=sheet_name)
                ws.append(col_labels)
                for row in rows:
                    ws.append([_cell_value(row.get(n)) for n in col_names])
        else:
            # **Single sheet** — historical behaviour, kept identical so existing screens.toml
            # files keep producing the same workbook layout.
            sheet_name = _safe_sheet_name(
                _substitute(sheet.name, ctx) or sheet.query, taken_sheet_names,
            )
            ws = wb.create_sheet(title=sheet_name)
            ws.append(col_labels)
            for row in result.rows:
                ws.append([_cell_value(row.get(n)) for n in col_names])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _split_values(
    screen: Screen,
    *,
    split_by: str,
    app: str,
    connectors: ConnectorRegistry,
    language: str | None,
    user: str | None,
) -> list[Any]:
    """Run the screen's read query and pull the distinct values of ``split_by`` (preserving
    first-seen order — so groups stay in the natural order the database returned them,
    matching what the operator sees in the TableView)."""
    conn_name = screen.connector or app
    try:
        conn = connectors.sql(conn_name)
    except ConnectorError as exc:
        raise http_for_connector_error(exc) from exc
    try:
        result = await conn.execute(
            screen.read_query, {},
            language=language, user=user,
            column_hints=screen.columns,
            screen_max_rows=screen.max_rows,
        )
    except ConnectorError as exc:
        raise http_for_connector_error(exc) from exc
    # Case-insensitive column-name match (Postgres folds unquoted identifiers; the operator
    # might type the column in any case in ``split_by``).
    target = split_by.upper()
    name_map = {c.name.upper(): c.name for c in result.columns}
    real_name = name_map.get(target)
    if real_name is None:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=f"split_by column {split_by!r} not in read_query result columns",
        )
    seen: set[Any] = set()
    out: list[Any] = []
    for r in result.rows:
        v = r.get(real_name)
        # Treat None / empty string as a single "no value" bucket (same as v1).
        key = v if (v is not None and v != "") else None
        if key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out


def _filename_for(template: str | None, *, screen: Screen, split_value: Any) -> str:
    """Per-workbook file name. ``{{split_value}}`` + ``{{screen}}`` placeholders.
    Default: ``<screen>_<split_value>.xlsx`` when bursting, ``<screen>.xlsx`` otherwise.
    Sanitises filesystem-unsafe characters (POSIX is loose, but Windows trips on more)."""
    if template:
        name = _substitute(template, {"split_value": split_value, "screen": screen.id})
    elif split_value is not None:
        name = f"{screen.id}_{split_value}"
    else:
        name = screen.id
    name = re.sub(r'[/\\:*?"<>|]+', "_", str(name)).strip(". ")
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return name


def _can_export(p: Principal, screen: Screen, *, app: str) -> bool:
    """The caller must hold the read perm for every query the export will fire. Same gating
    convention used by ``GET /api/screens`` and the action routes — surface the screen as
    404 (not 403) for a caller who's missing perms, so the screen's existence doesn't leak."""
    conn = screen.connector or app
    if not p.has_permission(f"sql:{conn}:{screen.read_query}"):
        return False
    if screen.export:
        for sheet in screen.export.sheets:
            sc = sheet.connector or conn
            if not p.has_permission(f"sql:{sc}:{sheet.query}"):
                return False
    return True


@router.post("/screens/{app}/{screen_id}/export", summary="Export screen to Excel")
async def export_workbook(
    app: str,
    screen_id: str,
    request: Request,
    principal: CurrentPrincipal,
    screens: Screens,
    connectors: Connectors,
) -> StreamingResponse:
    """Build the workbook(s) configured on ``screens.toml`` and stream the result.

    Body: empty (the export config + the running user are the only inputs). Future shape
    can grow per-request overrides (filter the split_value list, override the file name)
    via JSON body; for now the configuration on disk drives everything.

    Permission gate: the caller must hold ``sql:<conn>:<read_query>`` AND
    ``sql:<conn>:<query>`` for every sheet. Missing any of those → 404 (not 403) so the
    presence of the screen isn't leaked.
    """
    app_screens = screens.screens.get(app) or {}
    screen = app_screens.get(screen_id)
    if screen is None or not _can_export(principal, screen, app=app):
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"Unknown screen {app!r}/{screen_id!r}")
    spec = screen.export
    if spec is None or not spec.sheets:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=f"Screen {app}/{screen_id} has no [export] config (or no sheets).",
        )

    language = request_language(request)
    user = principal.username

    if spec.split_by:
        values = await _split_values(
            screen,
            split_by=spec.split_by,
            app=app, connectors=connectors,
            language=language, user=user,
        )
        if not values:
            # Empty list → emit a single xlsx for the "no group" bucket so the operator gets
            # *something* back instead of an empty zip. Less surprising than a 200 with empty
            # body. ``split_value`` stays ``None``; the sheets bind that.
            values = [None]
    else:
        values = [None]                                         # single-workbook mode

    files: list[tuple[str, bytes]] = []
    for v in values:
        xlsx = await _build_one_workbook(
            spec=spec, split_value=v,
            screen=screen, app=app, connectors=connectors,
            language=language, user=user,
        )
        files.append((_filename_for(spec.file_name_template, screen=screen, split_value=v), xlsx))

    # Single workbook → return raw .xlsx. Multiple → pack into a .zip + send. The
    # ``Content-Disposition`` filename hint drives the browser's downloaded name; the
    # response is streamed via ``StreamingResponse`` for predictable memory use.
    if len(files) == 1:
        name, content = files[0]
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{name}"'},
        )
    # Multi-workbook zip — name from ``archive_name`` or default ``<screen>.zip``.
    archive_name = spec.archive_name or f"{screen.id}.zip"
    if not archive_name.lower().endswith(".zip"):
        archive_name = f"{archive_name}.zip"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, content in files:
            zf.writestr(name, content)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{archive_name}"'},
    )
